from flask import Flask, render_template, request, send_file
import pandas as pd
import os
from openpyxl import load_workbook
from zipfile import BadZipFile

app = Flask(__name__)

UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'output'
TEMPLATE_FILE = 'template.xlsx'

# Create folders if they don’t exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

ALLOWED_EXTENSIONS = {'csv'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/')
def index():
    return '''
        <h2>Upload CSV File</h2>
        <form method="POST" action="/upload" enctype="multipart/form-data">
            <input type="file" name="csv_file" accept=".csv" required>
            <button type="submit">Upload</button>
        </form>
    '''

@app.route('/upload', methods=['POST'])
def upload_csv():
    if 'csv_file' not in request.files:
        return "❌ No file uploaded"

    file = request.files['csv_file']
    if file.filename == '':
        return "❌ No file selected"

    if not allowed_file(file.filename):
        return "❌ Only CSV files are allowed"

    filepath = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(filepath)

    try:
        if not os.path.exists(TEMPLATE_FILE):
            return "❌ Template file not found"

        # ✅ Read CSV using pandas with no header
        df = pd.read_csv(filepath, header=None)

        # ✅ Load Excel Template
        wb = load_workbook(TEMPLATE_FILE)
        ws = wb.active

        # ✅ A1 → B5
        if df.shape[0] >= 1:
            value_a1 = str(df.iloc[0, 0]).strip()
            ws['B5'] = value_a1
            print("✅ CSV A1 → Excel B5 =", value_a1)

        # ✅ A5 → B2
        if df.shape[0] >= 5:
            value_a5 = str(df.iloc[4, 0]).strip()
            ws['B2'] = value_a5
            print("✅ CSV A5 → Excel B2 =", value_a5)

        # ✅ A16 → B14
        if df.shape[0] >= 16:
            value_a16 = str(df.iloc[15, 0]).strip()
            ws['B14'] = value_a16
            print("✅ CSV A16 → Excel B14 =", value_a16)

        # ✅ A19 → B16
        if df.shape[0] >= 19:
            value_a19 = str(df.iloc[18, 0]).strip()
            ws['B16'] = value_a19
            print("✅ CSV A19 → Excel B16 =", value_a19)

        # ✅ Save and return filled Excel file
        output_file = f"filled_{os.path.splitext(file.filename)[0]}.xlsx"
        output_path = os.path.join(OUTPUT_FOLDER, output_file)
        wb.save(output_path)

        return send_file(output_path, as_attachment=True)

    except BadZipFile:
        return "❌ Error: Your template is not a valid .xlsx file. Please re-save using Excel."

    except Exception as e:
        return f"❌ Unexpected Error: {str(e)}"

if __name__ == '__main__':
    app.run(debug=True)
